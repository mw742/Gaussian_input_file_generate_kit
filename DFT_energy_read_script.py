
import openpyxl

## open an excel file to store the single point energy value from gaussian output files
data=openpyxl.load_workbook('all_mols_energy.xlsx')
## to store in the first page of energy excel file, sheetnames[0]
sheetnames=data.get_sheet_names()
table=data.get_sheet_by_name(sheetnames[0])
table=data.active
#print(table.title)


## to read the single point energy value from gaussian output files
## now have 118 output files to be handled
path="/users/wmm/desktop/rdkit_DFT/output/mw742/"
file_order=0
while file_order < 119:
    with open(path+str(file_order)+".out", "r") as f:
        res=f.readlines()
        for line in res:
            if line.find("SCF Done:  E(RB3LYP) ") > -1:
                energy=float(line[25:44])
                #print(energy)
                #print(type(energy))
                table.cell(file_order+1, 1).value=energy
    file_order+=1

data.save('all_mols_energy.xlsx')