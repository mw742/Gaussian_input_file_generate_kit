
##########################################################################
##
## Here is the script to read the results of Maestro job from .log file
##
##########################################################################



import openpyxl

## build an excel table to store all the parameters generated
## open an excel file to store the single point energy value from gaussian output files

## here to changen with working path
## the empty excel file should be created already
data=openpyxl.load_workbook('/users/wmm/desktop/Maestro_result.xlsx')
## to store in the first page of energy excel file, sheetnames[0]
sheetnames=data.get_sheet_names()
table=data.get_sheet_by_name(sheetnames[0])
table=data.active
#print(table.title)


## The results are save in seperate excel file
num=1
while num < 14:
    with open("/users/wmm/desktop/RESULTS/compare_"+str(num)+".log", 'r') as f:
        f_txt=f.read().splitlines()
        #print(type(f_txt))
        #print("f_txt:", f_txt)
        final_report_index=f_txt.index("Final report:")
        #print("final_report_index:", final_report_index)
        num_of_uni_stru=int(f_txt[final_report_index+1].split("unique")[0])
        print("num_of_uni_stru:", num_of_uni_stru)

        global_mini_line=[i for i in f_txt if "Conformation       1" in i]
        global_mini_index=f_txt.index(global_mini_line[0])
        #print("global_mini_index:", global_mini_index)
        num_of_global_mini=int(f_txt[global_mini_index].split("found")[1].split("times")[0])
        global_mini_energy=f_txt[global_mini_index].split("(")[1].split(")")[0]
        print("num_of_global_mini:", num_of_global_mini)
        print("global_mini_energy:", global_mini_energy)

        ## count the structures which are within 1 kJ/mol of the global minimum, what is the smallest number of times any of them has been found
        num_within_mini_range_index=final_report_index+3
        num_within_mini_range=int(f_txt[num_within_mini_range_index].split("confs")[0].split("Found")[1])
        a=0
        current_mini_num=0
        current_index=global_mini_index
        smallest_mini_num=num_of_global_mini
        while a < num_within_mini_range:
            current_mini_num=int(f_txt[current_index].split("found")[1].split("times")[0])
            if current_mini_num < smallest_mini_num:
                smallest_mini_num=current_mini_num
            else:
                pass
            current_index+=1
            a+=1
        print("smallest_mini_num:", smallest_mini_num)

        ## calucltae the unique structure found / total structures, if the value is smaller than 50% (experienced estimated),
        ## it might represent the cearch has been done
        total_structure_index=final_report_index-2
        total_structure_num=int(f_txt[total_structure_index].split("E")[0].split("Conf")[1])
        print("total_structure_num:", total_structure_num)
        experienced_para=num_of_uni_stru/total_structure_num
        print("experienced_para:", experienced_para)

        ## calculate the time to run a search
        time_line=[i for i in f_txt if "Time in Monte Carlo generation loop:" in i]
        time_index=f_txt.index(time_line[0])
        conformer_generate_time=float(f_txt[time_index].split(":")[1].split("CPU")[0])
        energy_minimization_time=float(f_txt[time_index+1].split(":")[1].split("CPU")[0])
        total_time=conformer_generate_time+energy_minimization_time
        print("total_time:", total_time)

        ## save all these data into excel table
        ## the first element in the bracket is row and the second is column
        table.cell(num+1, 2).value=num_of_uni_stru
        table.cell(num+1, 3).value=num_of_global_mini
        table.cell(num+1, 4).value=global_mini_energy
        table.cell(num+1, 5).value=num_within_mini_range
        table.cell(num+1, 6).value=smallest_mini_num
        table.cell(num+1, 7).value=experienced_para
        table.cell(num+1, 8).value=total_time

        if (num_of_global_mini > 2) and (experienced_para < 0.5):
            table.cell(num+1, 9).value="PASS"
        else:
            table.cell(num+1, 9).value="Undone"
    num+=1


## save the excel file in workig path
data.save('/users/wmm/desktop/Maestro_result.xlsx')
